# -*- coding:utf-8 -*-

import os
import time
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, colors
import pandas as pd
from atp.api.comm_log import logger
from atp.config.default import get_config

CONFIG = get_config()


class ExcelParser(object):

    def __init__(self, excel_name):
        self.ws_title = excel_name  # excel文件sheet名称
        self.border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
                             top=Side(style='thin', color='000000'),
                             bottom=Side(style='thin', color='000000'))  # excel表格样式
        self.font = Font(bold=True, size=10)  # excel表格标题栏样式
        self.fill = PatternFill("solid", fgColor="538DD5")  # excel表格标题栏样式
        self.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)  # 设置单元格样式

    def writeExcel(self, values):
        # 写excel文件
        wb = Workbook()
        #ws = wb.active
        #ws.title = self.ws_title
        if not os.path.exists(CONFIG.DOWNLOADS_DIR):
            os.makedirs(CONFIG.DOWNLOADS_DIR)
        file_name = "{0}_业务用例_{1}.xlsx".format(self.ws_title, int(time.time()))
        # file_name = "{0}_业务用例.xlsx".format(self.ws_title)
        file_dir = CONFIG.DOWNLOADS_DIR + file_name

        # 创建sheet
        sheet_idx=0
        for module_name, value in values.items():
            ws = wb.create_sheet(module_name, sheet_idx)
            if value:
                row_idx = 0
                for i in range(0, len(value)):
                    if i>1:
                        '''row_idx表示前面用例所占操作步骤总行数'''
                        if len(value[i-1][4])<=1:
                            row_idx=row_idx
                        else:
                            row_idx = row_idx+len(value[i-1][4])-1   #2
                    for j in range(0, len(value[i])):
                        if i == 0:
                            ws.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
                        elif i == 1:
                            if isinstance(value[i][j],list) :
                                if len(value[i][j]) >=1:
                                    for x in range(0,len(value[i][j])):
                                        ws.cell(row=i + 1 + x, column=j + 1, value=str(value[i][j][x]))
                                elif len(value[i][j])==0:
                                        ws.cell(row=i + 1  , column=j + 1, value=str(''))
                            else:
                                ws.cell(row=i + 1 , column=j + 1, value=str(value[i][j]))
                        else:
                            if isinstance(value[i][j],list) :
                                if len(value[i][j])>=1:
                                    for x in range(0,len(value[i][j])):
                                        ws.cell(row= i + 1 + row_idx+ x, column=j + 1, value=str(value[i][j][x]))
                                elif len(value[i][j])==0:
                                    ws.cell(row=i + 1 + row_idx, column=j + 1, value=str(''))
                            else:
                                ws.cell(row=i + 1 + row_idx, column=j + 1, value=str(value[i][j]))
                for cols in ws.iter_cols():
                    for cell in cols:
                        cell.font = Font(size=9)
                        cell.border = self.border
                        cell.alignment = self.alignment
                    cell_title = cols[0]
                    cell_title.font = self.font
                    cell_title.fill = self.fill
                ws.row_dimensions[1].height = 20
                ws.column_dimensions['A'].width = 10
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 40
                ws.column_dimensions['F'].width = 40
                ws.column_dimensions['G'].width = 20
                ws.column_dimensions['H'].width = 20
                ws.freeze_panes = 'A2'
            else:
                raise Exception("存储Excel失败")
            sheet_idx +=1
        wb.save(file_dir)
        logger.info("写入数据成功！")
        return file_name

    def write_summary_to_excel(self, run_date, value_lists):
        wb = Workbook()
        ws = wb.create_sheet(run_date)
        for value_list in value_lists:
            ws.append(value_list)
        if not os.path.exists(CONFIG.DOWNLOADS_DIR):
            os.makedirs(CONFIG.DOWNLOADS_DIR)
        excel_name = '{0}_每日测试任务结果_{1}.xlsx'.format(run_date, int(time.time()))
        ws.row_dimensions[1].height = 20
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 40
        ws.freeze_panes = 'A2'
        del wb['Sheet']
        wb.save(r'{0}{1}'.format(CONFIG.DOWNLOADS_DIR, excel_name))
        return excel_name

    def write_export_to_excel(self, start_time, value_lists):
        wb = Workbook()
        ws = wb.create_sheet('创建时间大于{}的用例'.format(start_time))
        for value_list in value_lists:
            ws.append(value_list)
        excel_name = '{0}_{1}.xlsx'.format(start_time, int(time.time()))
        ws.row_dimensions[1].height = 20
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.freeze_panes = 'A2'
        del wb['Sheet']
        excel_path = r'{0}{1}'.format(CONFIG.DOWNLOADS_DIR, excel_name)
        wb.save(excel_path)
        return excel_name

    def write_smoking_test_log_to_excel(self, test_type, value_list):
        if not value_list:
            return None
        df = pd.DataFrame(value_list[1:], columns=value_list[0])
        excel_name = '{0}_{1}.xlsx'.format(test_type, int(time.time()))
        excel_path = r'{0}{1}'.format(CONFIG.DOWNLOADS_DIR, excel_name)
        writer = pd.ExcelWriter(excel_path)
        df.to_excel(writer, 'Sheet1', index=False)
        writer.sheets['Sheet1'].column_dimensions['A'].width = 10
        writer.sheets['Sheet1'].column_dimensions['B'].width = 30
        writer.sheets['Sheet1'].column_dimensions['C'].width = 10
        writer.sheets['Sheet1'].column_dimensions['D'].width = 10
        writer.sheets['Sheet1'].column_dimensions['E'].width = 10
        writer.sheets['Sheet1'].column_dimensions['F'].width = 10
        writer.sheets['Sheet1'].column_dimensions['G'].width = 15
        writer.sheets['Sheet1'].column_dimensions['H'].width = 15
        writer.sheets['Sheet1'].column_dimensions['I'].width = 20
        writer.sheets['Sheet1'].column_dimensions['J'].width = 20
        writer.sheets['Sheet1'].column_dimensions['K'].width = 20
        writer.save()
        return excel_name

if __name__ == '__main__':
    value={
    "模块1":[
            ['编号' '模块', '用例标题', '前置条件', '操作步骤', '预期结果','执行结果'],
            [1,'模块11', '用例标题1', ['这是前置条件', '前置22', ''] ,['步骤1', '步骤2', '步骤3'], ['预期1', '预期2', '预期3']],
            [2,'模块22', '用例标题2', ['这是前置条件', '前置22'] ,['步骤1', '步骤2' ], ['预期1', '预期2']],
            [3,'模块33', '用例标题3', ['这是前置条件', '前置22', ''] ,['步骤1', '步骤2', '步骤3'], ['预期1', '预期2', '预期3']],
        [4, '放款订单领取列表_列表查询', '输入承租方查询，放款订单领取列表检索结果正确', [''],
         ['出纳进入放款订单领取列表，输入承租方，点击查询'],
         ['查询结果正确，放款订单领取列表只展示当前该承租方的数据，支持模糊查询，列表展示的订单编号、承租方、手机号、产品名称、用车对象、车辆类型、车辆信息、总融资额、期数、审核通过时间、放款金额和当前状态数据 正确'],
         [''],
         ['']
         ],
        [5, '放款订单领取列表_列表查询', '当前状态遍历所有状态值查询，放款订单领取列表检索结果正确', [''],
         ['出纳进入放款订单领取列表，当前状态遍历所有状态值，点击查询'],
         ['查询结果正确，放款订单领取列表只展示查询状态对应的数据，支持模糊查询，列表展示的订单编号、承租方、手机号、产品名称、用车对象、车辆类型、车辆信息、总融资额、期数、审核通过时间、放款金额和当前状态数据 正确'],
         [''],
         ['']
         ],
        [6, '放款订单领取列表_列表查询', '重置查询条件成功', [''],
         ['出纳进入放款订单领取列表，查询条件输入或选择数据后，点击重置'],
         ['查询条件重置成功'],
         [''],
         ['']
         ],
    ],
    "模块2":[
        [ '模块', '用例标题', '前置条件', '操作步骤', '预期结果', '执行结果', '备注'],
        [1, '模块111', '用例标题1', [''], ['步骤1', '步骤2', '步骤3'], ['']],
        [2, '模块222', '用例标题1', [''], [''], ['']],
        [3, '模块333', '用例标题3', ['这是前置条件', '前置22', ''], ['步骤1', '步骤2', '步骤3'], ['预期1', '预期2', '预期3']],

    ],
}
    ss = ExcelParser("测试")
    res = ss.writeExcel(value)
    print(res)
